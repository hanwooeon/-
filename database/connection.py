import psycopg2
import psycopg2.extras
import json
import hashlib
import re
from datetime import datetime
from pathlib import Path
import csv
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class DatabaseManager:
    def __init__(self, db_config: dict = None):
        if db_config is None:
            # 순환 import 방지를 위해 여기서 import
            from config.settings import DATABASE_CONFIG
            self.db_config = DATABASE_CONFIG
        else:
            self.db_config = db_config
        
        # 텍스트 파일 저장을 위한 디렉토리 생성
        self.results_dir = Path('results')
        self.results_dir.mkdir(exist_ok=True)
        
        self.init_database()
    
    def get_connection(self):
        return psycopg2.connect(**self.db_config)
    
    def init_database(self):
        """데이터베이스 테이블 초기화"""
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                
                # 키워드 테이블 생성
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS keywords (
                        id SERIAL PRIMARY KEY,
                        category VARCHAR(100) NOT NULL,
                        keyword TEXT NOT NULL,
                        UNIQUE(category, keyword)
                    )
                ''')
                
                # 크롤링 결과 테이블 생성
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS crawl_results (
                        id SERIAL PRIMARY KEY,
                        url TEXT NOT NULL UNIQUE,
                        title TEXT,
                        content TEXT,
                        content_hash TEXT NOT NULL,
                        detected_keywords JSONB,
                        is_processed BOOLEAN DEFAULT FALSE,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                ''')
                
                # 기존 테이블에 created_at 컬럼 추가 (이미 존재하면 무시)
                try:
                    cursor.execute('''
                        ALTER TABLE crawl_results 
                        ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    ''')
                except:
                    pass  # 이미 존재하거나 다른 이유로 실패해도 계속 진행
                
                # 내용 해시에 인덱스 추가
                cursor.execute('''
                    CREATE INDEX IF NOT EXISTS idx_crawl_results_content_hash 
                    ON crawl_results(content_hash)
                ''')
                
                conn.commit()
                print("PostgreSQL 데이터베이스 테이블 초기화 완료")
        except psycopg2.Error as e:
            print(f"데이터베이스 초기화 오류: {e}")
    
    def insert_keywords(self, keywords_data):
        """키워드 데이터 삽입"""
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                for category, keywords in keywords_data.items():
                    for keyword in keywords:
                        cursor.execute('''
                            INSERT INTO keywords (category, keyword)
                            VALUES (%s, %s)
                            ON CONFLICT (category, keyword) DO NOTHING
                        ''', (category, keyword))
                conn.commit()
                print("키워드 데이터 삽입 완료")
        except psycopg2.Error as e:
            print(f"키워드 삽입 오류: {e}")
    
    
    def get_keywords(self, category=None):
        """키워드 조회"""
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
                if category:
                    cursor.execute('SELECT * FROM keywords WHERE category = %s', (category,))
                else:
                    cursor.execute('SELECT * FROM keywords')
                return cursor.fetchall()
        except psycopg2.Error as e:
            print(f"키워드 조회 오류: {e}")
            return []
    
    def save_crawl_result(self, url, title, content, detected_keywords):
        """크롤링 결과 저장 (DB + 텍스트 파일) with duplicate prevention"""
        content_hash = self._get_content_hash(content)
        
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                
                # 1. URL 중복 체크
                cursor.execute('SELECT id, title FROM crawl_results WHERE url = %s', (url,))
                existing = cursor.fetchone()
                if existing:
                    return "duplicate_url"
                
                # 2. 내용 중복 체크 (해시 기반)
                cursor.execute('SELECT id, url, title FROM crawl_results WHERE content_hash = %s', (content_hash,))
                existing = cursor.fetchone()
                if existing:
                    return "duplicate_content"
                
                # 3. 새로운 내용이면 저장
                cursor.execute('''
                    INSERT INTO crawl_results (url, title, content, content_hash, detected_keywords)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id, created_at
                ''', (url, title, content, content_hash, psycopg2.extras.Json(detected_keywords)))
                conn.commit()
                result = cursor.fetchone()
                result_id, created_at = result[0], result[1]
                
                # 엑셀 파일에만 저장 (하이퍼링크 포함)
                if EXCEL_AVAILABLE:
                    self._save_to_excel_file(result_id, url, title, content, detected_keywords, created_at)
                else:
                    print("openpyxl을 설치할 수 없어 엑셀 저장이 비활성화되었습니다.")
                
                print(f"새로운 크롤링 결과 저장 완료 (ID: {result_id}) - Excel 파일에 저장됨")
                return result_id
                
        except psycopg2.Error as e:
            if "duplicate key value violates unique constraint" in str(e):
                print(f"중복 URL로 인한 저장 실패: {url}")
                return None
            print(f"크롤링 결과 저장 오류: {e}")
            return None
    
    def _normalize_content_for_duplicate_check(self, content: str) -> str:
        """중복 체크를 위한 내용 정규화"""
        if not content:
            return ""
        
        # 기본 정규화
        normalized = content.strip()
        
        # 앞에 붙는 노이즈 제거 (None, 공백 등)
        prefixes_to_remove = ['None', 'null', 'undefined', '']
        for prefix in prefixes_to_remove:
            if normalized.startswith(prefix):
                normalized = normalized[len(prefix):].strip()
        
        # 연속된 공백을 단일 공백으로
        normalized = re.sub(r'\s+', ' ', normalized)
        
        # 앞뒤 공백 제거
        normalized = normalized.strip()
        
        return normalized

    def _get_content_hash(self, content: str) -> str:
        """내용의 해시값 생성"""
        normalized_content = self._normalize_content_for_duplicate_check(content)
        return hashlib.md5(normalized_content.encode('utf-8')).hexdigest()


    def _is_duplicate_content(self, content: str) -> bool:
        """내용이 중복인지 확인 (DB만 사용)"""
        content_hash = self._get_content_hash(content)
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT id FROM crawl_results WHERE content_hash = %s', (content_hash,))
                return cursor.fetchone() is not None
        except psycopg2.Error as e:
            print(f"내용 중복 체크 오류: {e}")
            return False

    def _save_to_text_file(self, result_id, url, title, content, detected_keywords, created_at=None):
        """크롤링 결과를 JSON 및 테이블 형식 TXT 파일에 저장"""
        try:
            # JSON 파일 저장
            json_filename = "crawl_results.json"
            json_filepath = self.results_dir / json_filename
            
            # 새 결과 데이터
            timestamp_str = created_at.strftime('%Y-%m-%d %H:%M:%S') if created_at else datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            new_result = {
                "id": result_id,
                "timestamp": timestamp_str,
                "url": url,
                "title": title,
                "content": content,
                "detected_keywords": detected_keywords
            }
            
            # 기존 결과 로드
            results = []
            if json_filepath.exists():
                try:
                    with open(json_filepath, 'r', encoding='utf-8') as f:
                        results = json.load(f)
                except (json.JSONDecodeError, FileNotFoundError):
                    results = []
            
            # 새 결과 추가
            results.append(new_result)
            
            # JSON 파일에 저장
            with open(json_filepath, 'w', encoding='utf-8') as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            
            # 테이블 형식 TXT 파일 저장
            txt_filename = "crawl_results.txt"
            txt_filepath = self.results_dir / txt_filename
            
            content_hash = self._get_content_hash(content)
            
            # 헤더가 없으면 추가
            if not txt_filepath.exists():
                with open(txt_filepath, 'w', encoding='utf-8') as f:
                    # 고정 너비 컬럼 헤더 (엑셀 스타일)
                    f.write(f"{'ID':<5} | {'TIME':<19} | {'URL':<60} | {'TITLE':<30} | {'CONTENT':<50} | {'HASH':<12}\n")
                    f.write("-" * 200 + "\n")
            
            # 데이터 추가
            with open(txt_filepath, 'a', encoding='utf-8') as f:
                # 내용을 한 줄로 만들기 (줄바꿈 제거)
                clean_content = content.replace('\n', ' ').replace('\r', ' ').strip() if content else 'N/A'
                clean_title = title.replace('\n', ' ').replace('\r', ' ').strip() if title else 'N/A'
                
                # 엑셀 스타일: 고정 너비로 정렬하되 넘치면 잘리지만 전체 데이터는 표시
                # 각 필드를 고정 너비로 맞춤 (왼쪽 정렬)
                id_str = f"{result_id}"[:5].ljust(5)
                time_str = f"{new_result['timestamp']}"[:19].ljust(19)
                url_str = f"{url}"[:60].ljust(60) if len(url) <= 60 else f"{url[:57]}..."
                title_str = f"{clean_title}"[:30].ljust(30) if len(clean_title) <= 30 else f"{clean_title[:27]}..."
                content_str = f"{clean_content}"[:50].ljust(50) if len(clean_content) <= 50 else f"{clean_content[:47]}..."
                hash_str = f"{content_hash[:12]}"
                
                # 전체 URL을 주석으로 추가 (긴 URL의 경우)
                if len(url) > 60:
                    f.write(f"{id_str} | {time_str} | {url_str} | {title_str} | {content_str} | {hash_str}\n")
                    f.write(f"{'':>5} | {'':>19} | FULL_URL: {url}\n")
                else:
                    f.write(f"{id_str} | {time_str} | {url_str} | {title_str} | {content_str} | {hash_str}\n")
            
            print(f"JSON 및 테이블 형식 TXT 파일에 저장: {json_filepath}, {txt_filepath}")
            
        except Exception as e:
            print(f"파일 저장 오류: {e}")
    
    def _save_to_csv_file(self, result_id, url, title, content, detected_keywords, created_at=None):
        """크롤링 결과를 CSV 파일에 저장 (엑셀에서 열기 가능)"""
        try:
            csv_filename = "crawl_results.csv"
            csv_filepath = self.results_dir / csv_filename
            
            timestamp_str = created_at.strftime('%Y-%m-%d %H:%M:%S') if created_at else datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # 내용 정리 (줄바꿈, 쉼표 등 CSV 문제가 될 수 있는 문자 처리)
            clean_content = content.replace('\n', ' ').replace('\r', ' ').replace('"', '""').strip() if content else 'N/A'
            clean_title = title.replace('\n', ' ').replace('\r', ' ').replace('"', '""').strip() if title else 'N/A'
            
            # 키워드 정보를 간단한 텍스트로 변환
            keyword_summary = ""
            if detected_keywords:
                for category, keywords in detected_keywords.items():
                    if keywords:
                        keyword_list = [kw.get('keyword', '') for kw in keywords]
                        keyword_summary += f"{category}: {', '.join(set(keyword_list))}; "
            keyword_summary = keyword_summary.strip('; ')
            
            content_hash = self._get_content_hash(content)
            
            # 파일이 없으면 헤더와 함께 생성
            file_exists = csv_filepath.exists()
            
            with open(csv_filepath, 'a', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['ID', 'Timestamp', 'URL', 'Title', 'Content', 'Keywords', 'Hash']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                # 헤더 쓰기 (파일이 새로 생성된 경우)
                if not file_exists:
                    writer.writeheader()
                
                # 데이터 쓰기
                writer.writerow({
                    'ID': result_id,
                    'Timestamp': timestamp_str,
                    'URL': url,  # 전체 URL 저장
                    'Title': clean_title,
                    'Content': clean_content,
                    'Keywords': keyword_summary,
                    'Hash': content_hash[:12]
                })
            
            print(f"CSV 파일에 저장: {csv_filepath} (엑셀에서 열기 가능)")
            
        except Exception as e:
            print(f"CSV 파일 저장 오류: {e}")
    
    def _save_to_excel_file(self, result_id, url, title, content, detected_keywords, created_at=None):
        """크롤링 결과를 진짜 엑셀 파일(.xlsx)에 저장 (클릭 가능한 하이퍼링크 포함)"""
        try:
            excel_filename = "crawl_results.xlsx"
            excel_filepath = self.results_dir / excel_filename
            
            timestamp_str = created_at.strftime('%Y-%m-%d %H:%M:%S') if created_at else datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # 기존 파일이 있으면 로드, 없으면 새로 생성
            if excel_filepath.exists():
                wb = load_workbook(excel_filepath)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Crawl Results"
                
                # 헤더 스타일 설정
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                # 헤더 추가
                headers = ['ID', 'Timestamp', 'URL', 'Title', 'Content', 'Keywords', 'Hash']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 컬럼 너비 설정 (더 넓게)
                ws.column_dimensions['A'].width = 8   # ID
                ws.column_dimensions['B'].width = 20  # Timestamp  
                ws.column_dimensions['C'].width = 80  # URL (더 넓게)
                ws.column_dimensions['D'].width = 50  # Title
                ws.column_dimensions['E'].width = 80  # Content
                ws.column_dimensions['F'].width = 40  # Keywords
                ws.column_dimensions['G'].width = 15  # Hash
                
                # 헤더 행 높이 설정
                ws.row_dimensions[1].height = 25
            
            # 새 행 추가
            new_row = ws.max_row + 1
            
            # 내용 정리
            clean_content = content.replace('\n', ' ').replace('\r', ' ').strip() if content else 'N/A'
            clean_title = title.replace('\n', ' ').replace('\r', ' ').strip() if title else 'N/A'
            
            # 키워드 정보를 보기 좋게 변환
            keyword_summary = ""
            if detected_keywords:
                for category, keywords in detected_keywords.items():
                    if keywords:
                        unique_keywords = list(set([kw.get('keyword', '') for kw in keywords]))
                        keyword_summary += f"[{category}]: {', '.join(unique_keywords)}\n"
            keyword_summary = keyword_summary.strip()
            
            content_hash = self._get_content_hash(content)
            
            # 데이터 추가
            ws.cell(row=new_row, column=1, value=result_id)
            ws.cell(row=new_row, column=2, value=timestamp_str)
            
            # URL을 클릭 가능한 하이퍼링크로 추가
            url_cell = ws.cell(row=new_row, column=3, value=url)
            url_cell.hyperlink = url
            url_cell.font = Font(color="0563C1", underline="single")  # 파란색 밑줄
            
            ws.cell(row=new_row, column=4, value=clean_title)
            ws.cell(row=new_row, column=5, value=clean_content)
            ws.cell(row=new_row, column=6, value=keyword_summary)
            ws.cell(row=new_row, column=7, value=content_hash[:12])
            
            # 행 높이 설정 (긴 내용을 위해)
            ws.row_dimensions[new_row].height = 30
            
            # 텍스트 래핑 설정
            for col in range(1, 8):
                cell = ws.cell(row=new_row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # 파일 저장
            wb.save(excel_filepath)
            print(f"진짜 엑셀 파일(.xlsx)에 저장: {excel_filepath} (클릭 가능한 하이퍼링크 포함)")
            
        except Exception as e:
            print(f"엑셀 파일 저장 오류: {e}")
    
    def clean_duplicate_results(self):
        """기존 테이블 형식 결과 파일에서 중복 제거"""
        filepath = self.results_dir / "crawl_results.txt"
        if not filepath.exists():
            print("결과 파일이 존재하지 않습니다.")
            return
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            # 헤더와 구분선 찾기
            header_lines = []
            data_lines = []
            
            for i, line in enumerate(lines):
                if line.strip().startswith("ID | TIME") or line.strip().startswith("-" * 50):  # 구분선 길이 수정
                    header_lines.append(line)
                elif "|" in line and line.strip():
                    data_lines.append(line)
            
            seen_hashes = set()
            unique_data_lines = []
            duplicates_count = 0
            
            for line in data_lines:
                parts = line.split(" | ")
                if len(parts) >= 6:
                    # 해시값은 마지막 컬럼에서 추출
                    hash_part = parts[-1].strip()
                    if hash_part not in seen_hashes:
                        seen_hashes.add(hash_part)
                        unique_data_lines.append(line)
                    else:
                        duplicates_count += 1
            
            # 중복 제거된 내용으로 파일 재작성
            with open(filepath, 'w', encoding='utf-8') as f:
                # 헤더 먼저 쓰기
                for header_line in header_lines:
                    f.write(header_line)
                # 데이터 쓰기
                for data_line in unique_data_lines:
                    f.write(data_line)
            
            print(f"중복 제거 완료: {duplicates_count}개 중복 항목 제거됨")
            print(f"남은 고유 항목: {len(unique_data_lines)}개")
            
        except Exception as e:
            print(f"중복 제거 오류: {e}")
    
    def reset_database(self):
        """데이터베이스 완전 초기화 (모든 데이터 삭제)"""
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('DROP TABLE IF EXISTS crawl_results CASCADE')
                cursor.execute('DROP TABLE IF EXISTS keywords CASCADE')
                conn.commit()
                print("기존 테이블 삭제 완료")
                
            # 테이블 재생성
            self.init_database()
            print("데이터베이스 초기화 완료")
            
        except psycopg2.Error as e:
            print(f"데이터베이스 초기화 오류: {e}")
    
    def _is_duplicate_url(self, url: str) -> bool:
        """URL이 이미 존재하는지 확인"""
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT id FROM crawl_results WHERE url = %s', (url,))
                return cursor.fetchone() is not None
        except psycopg2.Error as e:
            print(f"URL 중복 체크 오류: {e}")
            return False